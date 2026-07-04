"""央企专区(CENTRAL_SOE)材料表 Excel 全量导入(Plan 2)。

把客户材料表(17 大类、去重 1646 SPU、含规格变体/单位)导入平台商品模型 + 专区白名单:
- 1 个 zone(CENTRAL_SOE,复用 demo seed 建的;无则建)
- 17 个 zone_categories(code=大类编号 01-17,name=大类名),幂等 upsert(改名复用 demo 的 01-05)
- 1646 个 products:visibility=ZONE_ONLY + status=ACTIVE,category_code 取自 master_final.final_code
  (1317 挂真实平台 leaf + 329 用 parent_hint 现有父 code 占位;code 是可晚绑定死元数据)
- SKU:多规格→每规格 1 个 ACTIVE SKU + selectable 属性 spec;单/无规格→1 个默认 SKU
- unit 取自 Excel 计量单位列(同 SPU 多单位取众数)
- 每 product 1 条 zone_product(挂对应 zone_category,source=IMPORT + source_batch_id)
- 清掉 demo 占位商品(spu_code LIKE 'ZONE-DEMO-%')+ 其 zone_products(保留 demo 买家 + zone_grant)
- NEW占位类目:上线前自动在 parent_hint 下创建真实叶子,商品最终只挂 leaf code

数据源(相对仓库根):
- data/zone_import/prepared/master_final.json —— SPU→final_code 决策(1646,0 校验错)
- data/zone_import/prepared/new_leaves.json —— NEW占位待建叶英文名/样例
- data/基础校准拓展:材料名称中英文、规格、单位.xlsx —— en/规格/单位(3724 行,join key=(大类,中文名))

join:master_final(大类,zh) ⋈ Excel(类别,中文名),用 NFKC+去空白 硬归一化(实测 1646↔1646 全中)。

幂等(可重复执行):
- zone: code;zone_category: (zone_id, code) 改名 upsert
- product: spu_code(deleted_at IS NULL);已存在则更新可变字段、不动 SKU(避免破坏 cart/rfq 的 sku_id 引用)
- zone_product: (zone_id, spu_id, zone_category_id)
- category leaf: (parent_code, name_zh) 复用;否则按父节点现有最大子 code + 1 生成
- spu_code/sku_code 由 (大类,中文名)/规格 的稳定 md5 派生,重跑不漂移

图片:本轮不导(v1 占位),前端走占位图 fallback;157 有 OVH 真图后续独立回填。

用法(本地/联调;生产严禁——脚本硬拦 OVH 生产库):
    cd backend
    # 先干跑看计划(不落库)
    python scripts/import_zone_soe_materials.py --dry-run
    # 确认后落库
    python scripts/import_zone_soe_materials.py --commit
"""
from __future__ import annotations

import argparse
import asyncio
import collections
import hashlib
import os
import re
import sys
import unicodedata
from pathlib import Path

# 允许 `python scripts/import_zone_soe_materials.py` 直接运行
_BACKEND_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_BACKEND_ROOT))
_REPO_ROOT = _BACKEND_ROOT.parent

import json  # noqa: E402

import openpyxl  # noqa: E402
from sqlalchemy import delete, func, select, update  # noqa: E402
from sqlalchemy.ext.asyncio import AsyncSession  # noqa: E402

from app.db.models.category import Category  # noqa: E402
from app.db.models.product import Product, ProductStatus, ProductVisibility  # noqa: E402
from app.db.models.product_attr import ProductAttr  # noqa: E402
from app.db.models.product_sku import ProductSku, SkuStatus  # noqa: E402
from app.db.models.zone import Zone, ZoneCategory, ZoneProduct  # noqa: E402

ZONE_CODE = "CENTRAL_SOE"
SOURCE = "IMPORT"
BATCH_ID = "zone_soe_material_v1"
DEMO_SPU_PREFIX = "ZONE-DEMO-"

MASTER_FINAL = _REPO_ROOT / "data" / "zone_import" / "prepared" / "master_final.json"
NEW_LEAVES = _REPO_ROOT / "data" / "zone_import" / "prepared" / "new_leaves.json"

# 17 大类编号 → (中文名, 英文名)。中文名须与 Excel/master_final 的大类名一致(join 与展示都靠它)。
ZONE_CATEGORIES = [
    ("01", "钢筋类", "Rebar"),
    ("02", "水泥类", "Cement"),
    ("03", "钢材类", "Steel"),
    ("04", "木材类", "Timber"),
    ("05", "成品砼类", "Ready-mix Concrete"),
    ("06", "砂石料类", "Aggregates"),
    ("07", "油漆涂料", "Paints & Coatings"),
    ("08", "电动工具", "Power Tools"),
    ("09", "平立面装修类", "Finishing Materials"),
    ("10", "五金耗材类", "Hardware & Consumables"),
    ("11", "给排水类", "Water Supply & Drainage"),
    ("12", "强弱电类", "Electrical"),
    ("13", "暖通气动类", "HVAC & Pneumatic"),
    ("14", "消防类", "Fire Protection"),
    ("15", "家具家电类", "Furniture & Appliances"),
    ("16", "周转料", "Reusable Materials"),
    ("17", "临时设施", "Temporary Facilities"),
]
_NAME_TO_CODE = {name_zh: code for code, name_zh, _ in ZONE_CATEGORIES}


def _hard(s: str) -> str:
    """硬归一化:NFKC(全角→半角)+ 去所有空白。用于 join key 与稳定 code 派生。"""
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(s)))


def _norm(v) -> str:
    return str(v).strip() if v not in (None, "") else ""


def _spu_code(cat_code: str, name_zh: str) -> str:
    h = hashlib.md5(_hard(name_zh).encode("utf-8")).hexdigest()[:8]
    return f"ZSOE-{cat_code}-{h}"


def _sku_code(spu_code: str, spec: str | None) -> str:
    if not spec:
        return f"{spu_code}-D"
    return f"{spu_code}-{hashlib.md5(_hard(spec).encode('utf-8')).hexdigest()[:6]}"


def _new_leaf_en_map() -> dict[tuple[str, str], str]:
    if not NEW_LEAVES.exists():
        return {}
    items = json.loads(NEW_LEAVES.read_text(encoding="utf-8"))
    return {
        (str(i["parent_hint_code"]), str(i["new_zh"])): str(i.get("new_en") or "")
        for i in items
    }


def _find_source_excel() -> Path:
    for p in (_REPO_ROOT / "data").glob("*.xlsx"):
        if "基础校准" in p.name:
            return p
    raise FileNotFoundError("找不到源材料表 Excel(data/基础校准拓展*.xlsx)")


def load_rows() -> list[dict]:
    """解析 Excel + master_final,join 出每 SPU 的导入行(SPU 级 + 变体规格)。"""
    excel_path = _find_source_excel()
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    # (hard(大类), hard(zh)) -> {en:[], specs:[], units:[]}
    ex: dict[tuple[str, str], dict] = {}
    order: list[tuple[str, str]] = []
    for sh in wb.sheetnames:
        if not sh.isdigit():
            continue
        for r, row in enumerate(wb[sh].iter_rows(values_only=True)):
            if r == 0:
                continue
            cat, zh, spec, unit = (_norm(row[i]) for i in (2, 3, 5, 6))  # 2类别 3中文名 5规格 6单位
            en = _norm(row[4])  # 4英文名
            if not zh:
                continue
            k = (_hard(cat), _hard(zh))
            d = ex.get(k)
            if d is None:
                d = {"en": [], "specs": [], "units": []}
                ex[k] = d
                order.append(k)
            if en:
                d["en"].append(en)
            if spec:
                d["specs"].append(spec)
            if unit:
                d["units"].append(unit)
    wb.close()

    master = json.loads(MASTER_FINAL.read_text(encoding="utf-8"))
    rows: list[dict] = []
    missing = []
    for m in master:
        cat_name, zh = m["大类"], m["zh"]
        cat_code = _NAME_TO_CODE.get(cat_name)
        if cat_code is None:
            raise RuntimeError(f"master_final 大类名不在 17 映射内:{cat_name!r}")
        d = ex.get((_hard(cat_name), _hard(zh)))
        if d is None:
            missing.append((cat_name, zh))
            continue
        # 去重规格(保序);单位取众数;英文名取第一个非空
        seen = set()
        specs = []
        for s in d["specs"]:
            if s not in seen:
                seen.add(s)
                specs.append(s)
        unit = collections.Counter(d["units"]).most_common(1)[0][0] if d["units"] else "PCS"
        name_en = d["en"][0] if d["en"] else None
        rows.append({
            "cat_code": cat_code,
            "cat_name": cat_name,
            "zh": zh,
            "en": name_en,
            "category_code": m["final_code"],
            "kind": m["kind"],
            "new_zh": m.get("new_zh") or "",
            "unit": unit,
            "specs": specs,  # 去重后规格串列表(可空)
            "spu_code": _spu_code(cat_code, zh),
        })
    if missing:
        raise RuntimeError(f"join 失败 {len(missing)} 条(master_final 有、Excel 无):{missing[:5]}")
    return rows


def _next_child_code(parent_code: str, child_codes: list[str]) -> str:
    prefix = f"{parent_code}."
    max_seg = 0
    for code in child_codes:
        if not code.startswith(prefix):
            continue
        rest = code[len(prefix):]
        if "." in rest or not rest.isdigit():
            continue
        max_seg = max(max_seg, int(rest))
    next_seg = max_seg + 1
    if next_seg > 999:
        raise RuntimeError(f"parent {parent_code} 子 code 已超过 999,无法自动建叶")
    code = f"{prefix}{next_seg:03d}"
    if len(code) > 16:
        raise RuntimeError(f"生成 code 超长({len(code)}): {code}")
    return code


async def _finalize_new_leaf_categories(db: AsyncSession, rows: list[dict], stats: dict) -> None:
    """把 NEW占位 parent_hint 自动收口成真实叶子,并回写 rows[*].category_code。"""
    specs: dict[tuple[str, str], str] = {}
    en_map = _new_leaf_en_map()
    for r in rows:
        if r["kind"] != "NEW占位":
            continue
        parent_code = r["category_code"]
        name_zh = r["new_zh"] or r["zh"]
        specs.setdefault((parent_code, name_zh), en_map.get((parent_code, name_zh), ""))

    if not specs:
        return

    parent_codes = sorted({p for p, _ in specs})
    parents = (await db.execute(
        select(Category).where(Category.code.in_(parent_codes), Category.is_active.is_(True))
    )).scalars().all()
    parent_by_code = {p.code: p for p in parents}
    missing = [p for p in parent_codes if p not in parent_by_code]
    if missing:
        raise RuntimeError(f"NEW占位 parent_hint 不存在:{missing[:10]}")

    existing_children = (await db.execute(
        select(Category).where(Category.parent_code.in_(parent_codes), Category.is_active.is_(True))
    )).scalars().all()
    children_by_parent: dict[str, list[Category]] = collections.defaultdict(list)
    child_by_parent_name: dict[tuple[str, str], Category] = {}
    for child in existing_children:
        children_by_parent[child.parent_code].append(child)
        child_by_parent_name[(child.parent_code, child.name_zh)] = child

    resolved: dict[tuple[str, str], Category] = {}
    for parent_code, name_zh in sorted(specs):
        existing = child_by_parent_name.get((parent_code, name_zh))
        if existing is not None:
            existing.is_leaf = True
            resolved[(parent_code, name_zh)] = existing
            stats["new_leaf_categories_existing"] += 1
            continue

        parent = parent_by_code[parent_code]
        child_code = _next_child_code(parent_code, [c.code for c in children_by_parent[parent_code]])
        child = Category(
            code=child_code,
            name_zh=name_zh,
            name_en=specs[(parent_code, name_zh)] or None,
            level=parent.level + 1,
            parent_code=parent_code,
            sort_order=max([c.sort_order for c in children_by_parent[parent_code]] or [0]) + 1,
            is_active=True,
            is_leaf=True,
            source_lang="zh",
            trans_meta={"name_zh": "src", "name_en": "manual" if specs[(parent_code, name_zh)] else "pending"},
        )
        db.add(child)
        await db.flush()
        children_by_parent[parent_code].append(child)
        child_by_parent_name[(parent_code, name_zh)] = child
        resolved[(parent_code, name_zh)] = child
        stats["new_leaf_categories_created"] += 1

    for parent in parent_by_code.values():
        if parent.is_leaf:
            parent.is_leaf = False
            stats["new_leaf_parent_marked_nonleaf"] += 1

    for r in rows:
        if r["kind"] != "NEW占位":
            continue
        parent_code = r["category_code"]
        name_zh = r["new_zh"] or r["zh"]
        r["parent_hint_code"] = parent_code
        r["category_code"] = resolved[(parent_code, name_zh)].code
        r["kind"] = "NEW叶子"
    await db.flush()


async def _validate_category_codes(db: AsyncSession, rows: list[dict]) -> dict:
    """校验所有 category_code 存在且为 active leaf。返回叶子 code 统计。"""
    codes = sorted({r["category_code"] for r in rows})
    found = (await db.execute(
        select(Category.code, Category.is_leaf).where(
            Category.code.in_(codes), Category.is_active.is_(True)
        )
    )).all()
    exist = {c: leaf for c, leaf in found}
    missing = [c for c in codes if c not in exist]
    if missing:
        raise RuntimeError(f"category_code 不存在于 categories,FK 会失败:{missing[:10]}")
    nonleaf = [c for c in codes if not exist[c]]
    if nonleaf:
        raise RuntimeError(f"category_code 仍有非叶子,上线前必须收口:{nonleaf[:10]}")
    return {
        "distinct_codes": len(codes),
        "leaf": len(codes),
        "nonleaf": 0,
    }


async def _get_or_create_zone(db: AsyncSession) -> Zone:
    zone = (await db.execute(select(Zone).where(Zone.code == ZONE_CODE))).scalar_one_or_none()
    if zone is not None:
        return zone
    zone = Zone(code=ZONE_CODE, name_zh="常用材料", name_en="Common Materials", name_sw="Vifaa vya Kawaida", status="ACTIVE")
    db.add(zone)
    await db.flush()
    return zone


async def _upsert_zone_categories(db: AsyncSession, zone: Zone, rows: list[dict], stats: dict) -> dict[str, ZoneCategory]:
    # mapped_platform_codes:各大类下 SPU 用到的 distinct category_code(选品辅助)
    mapped: dict[str, set] = collections.defaultdict(set)
    for r in rows:
        mapped[r["cat_code"]].add(r["category_code"])
    out: dict[str, ZoneCategory] = {}
    for i, (code, name_zh, name_en) in enumerate(ZONE_CATEGORIES):
        codes = sorted(mapped.get(code, set()))
        zc = (await db.execute(
            select(ZoneCategory).where(ZoneCategory.zone_id == zone.id, ZoneCategory.code == code)
        )).scalar_one_or_none()
        if zc is None:
            zc = ZoneCategory(
                zone_id=zone.id, code=code, name_zh=name_zh, name_en=name_en,
                sort_order=i, mapped_platform_codes=codes,
            )
            db.add(zc)
            stats["zone_categories_created"] += 1
        else:
            # 改名复用(demo 的 01-05 名字须纠正为真实大类名)+ 刷新映射
            zc.name_zh, zc.name_en, zc.sort_order, zc.mapped_platform_codes = name_zh, name_en, i, codes
            stats["zone_categories_updated"] += 1
        await db.flush()
        out[code] = zc
    return out


async def _purge_demo(db: AsyncSession, zone: Zone, stats: dict) -> None:
    """清 demo 占位商品 + 其 zone_products(保留 demo 买家/grant)。"""
    demo_ids = [pid for (pid,) in (await db.execute(
        select(Product.id).where(
            Product.spu_code.like(f"{DEMO_SPU_PREFIX}%"), Product.deleted_at.is_(None)
        )
    )).all()]
    if not demo_ids:
        return
    zp_deleted = (await db.execute(
        delete(ZoneProduct).where(ZoneProduct.zone_id == zone.id, ZoneProduct.spu_id.in_(demo_ids))
    )).rowcount
    # 软删 demo products(释放 spu_code 唯一约束;SKU/attr 随 cascade 保留但商品不可见)
    await db.execute(
        update(Product).where(Product.id.in_(demo_ids)).values(deleted_at=func.now())
    )
    stats["demo_products_purged"] = len(demo_ids)
    stats["demo_zone_products_purged"] = zp_deleted


async def _upsert_product(db: AsyncSession, r: dict, stats: dict) -> Product:
    spu_code = r["spu_code"]
    product = (await db.execute(
        select(Product).where(Product.spu_code == spu_code, Product.deleted_at.is_(None))
    )).scalar_one_or_none()
    if product is not None:
        # 已存在:更新可变字段,不动 SKU(避免破坏已下单引用的 sku_id)
        product.name_zh = r["zh"]
        product.name_en = r["en"]
        product.category_code = r["category_code"]
        product.moq_unit = r["unit"]  # 真实计量单位存 moq_unit(平台惯例:unit 恒 PCS)
        product.visibility = ProductVisibility.ZONE_ONLY
        product.status = ProductStatus.ACTIVE
        stats["products_updated"] += 1
        await db.flush()
        return product

    product = Product(
        spu_code=spu_code, name_zh=r["zh"], name_en=r["en"], category_code=r["category_code"],
        status=ProductStatus.ACTIVE, visibility=ProductVisibility.ZONE_ONLY,
        # 平台惯例:unit 恒默认 PCS(定价/下单读它),真实计量单位放 moq_unit
        unit="PCS", moq=1, moq_unit=r["unit"], source=SOURCE,
        source_meta={"batch": BATCH_ID, "大类": r["cat_name"], "kind": r["kind"]},
    )
    db.add(product)
    await db.flush()

    specs = r["specs"]
    if len(specs) >= 2:
        stats["products_variant"] += 1
        for i, spec in enumerate(specs):
            sku = ProductSku(
                product_id=product.id, sku_code=_sku_code(spu_code, spec),
                name_zh=f"{r['zh']} {spec}", name_en=None,
                moq=1, is_default=(i == 0), status=SkuStatus.ACTIVE,
            )
            db.add(sku)
            await db.flush()
            # SKU 级属性:详情页 SkuSelector(extractDimensions 扫 skus)渲染规格轴 + resolve_purchase_target 解析唯一 SKU
            db.add(ProductAttr(
                product_id=product.id, sku_id=sku.id,
                attr_key_zh="规格", attr_key_en="spec",
                attr_value_zh=spec, attr_value_en=spec,  # 规格值语言中性,en 同值以稳跨 locale 变体匹配
                selectable=True, sort_order=i,
            ))
        stats["skus_created"] += len(specs)
    else:
        stats["products_simple"] += 1
        db.add(ProductSku(
            product_id=product.id, sku_code=_sku_code(spu_code, None),
            name_zh=r["zh"], name_en=r["en"],
            moq=1, is_default=True, status=SkuStatus.ACTIVE,
        ))
        stats["skus_created"] += 1
        # 单规格(仅 1 个规格值)记为 SPU 级非可选属性,便于展示
        if len(specs) == 1:
            db.add(ProductAttr(
                product_id=product.id, sku_id=None,
                attr_key_zh="规格", attr_key_en="spec",
                attr_value_zh=specs[0], attr_value_en=specs[0],
                selectable=False, sort_order=0,
            ))
    await db.flush()
    return product


async def _upsert_zone_product(db: AsyncSession, zone: Zone, zc: ZoneCategory, product: Product, sort_order: int, stats: dict) -> None:
    zp = (await db.execute(
        select(ZoneProduct).where(
            ZoneProduct.zone_id == zone.id,
            ZoneProduct.spu_id == product.id,
            ZoneProduct.zone_category_id == zc.id,
        )
    )).scalar_one_or_none()
    if zp is not None:
        if zp.sort_order != sort_order or zp.source != SOURCE or zp.source_batch_id != BATCH_ID:
            zp.sort_order = sort_order
            zp.source = SOURCE
            zp.source_batch_id = BATCH_ID
            stats["zone_products_updated"] += 1
            await db.flush()
        else:
            stats["zone_products_existing"] += 1
        return
    db.add(ZoneProduct(
        zone_id=zone.id, spu_id=product.id, zone_category_id=zc.id,
        sort_order=sort_order, source=SOURCE, source_batch_id=BATCH_ID,
    ))
    stats["zone_products_created"] += 1
    await db.flush()


async def run_import(db: AsyncSession) -> dict:
    stats = collections.Counter()
    rows = load_rows()
    stats["spu_total"] = len(rows)
    await _finalize_new_leaf_categories(db, rows, stats)
    cat_stats = await _validate_category_codes(db, rows)

    zone = await _get_or_create_zone(db)
    await _purge_demo(db, zone, stats)
    zone_cats = await _upsert_zone_categories(db, zone, rows, stats)

    sort_counters: dict[str, int] = collections.defaultdict(int)
    for r in rows:
        product = await _upsert_product(db, r, stats)
        zc = zone_cats[r["cat_code"]]
        await _upsert_zone_product(db, zone, zc, product, sort_counters[r["cat_code"]], stats)
        sort_counters[r["cat_code"]] += 1

    return {"stats": dict(stats), "categories": cat_stats}


async def _execute(dry_run: bool) -> None:
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine
    from app.core.config import settings

    if (
        not dry_run
        and "162.19.98.142" in (settings.DATABASE_URL or "")
        and os.environ.get("ALLOW_ZONE_IMPORT_PROD") != "1"
    ):
        raise SystemExit(
            "拒绝执行:DATABASE_URL 指向 OVH 生产库。生产落库前请先完成备份并 dry-run, "
            "确认后设置 ALLOW_ZONE_IMPORT_PROD=1 再执行 --commit。"
        )

    engine = create_async_engine(settings.DATABASE_URL)
    Session = async_sessionmaker(engine, expire_on_commit=False)
    async with Session() as db:
        result = await run_import(db)
        if dry_run:
            await db.rollback()
        else:
            await db.commit()
    await engine.dispose()

    s = result["stats"]
    c = result["categories"]
    print("=" * 64)
    print(f"央企专区材料导入 {'[DRY-RUN 未落库]' if dry_run else '[已落库 COMMIT]'}  batch={BATCH_ID}")
    print("=" * 64)
    print(f"SPU 总数(join 后):       {s.get('spu_total', 0)}")
    print(f"category_code:            {c['distinct_codes']} distinct(leaf {c['leaf']} / 非叶 {c['nonleaf']})")
    print(f"NEW占位建叶:              created {s.get('new_leaf_categories_created',0)} / existing {s.get('new_leaf_categories_existing',0)} / parent_nonleaf {s.get('new_leaf_parent_marked_nonleaf',0)}")
    print(f"zone_categories:          created {s.get('zone_categories_created',0)} / updated {s.get('zone_categories_updated',0)}")
    print(f"demo 清理:                products {s.get('demo_products_purged',0)} / zone_products {s.get('demo_zone_products_purged',0)}")
    print(f"products 新建:            simple {s.get('products_simple',0)} + variant {s.get('products_variant',0)} = {s.get('products_simple',0)+s.get('products_variant',0)}")
    print(f"products 更新(已存在):    {s.get('products_updated',0)}")
    print(f"SKU 新建:                 {s.get('skus_created',0)}")
    print(f"zone_products:            created {s.get('zone_products_created',0)} / updated {s.get('zone_products_updated',0)} / existing {s.get('zone_products_existing',0)}")
    print("注:图片本轮未导(v1 占位),前端走占位图 fallback;后续独立回填。")
    if dry_run:
        print("\n这是干跑,数据库未改动。确认无误后加 --commit 落库。")


def main() -> None:
    ap = argparse.ArgumentParser(description="央企专区材料表 Excel 全量导入")
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--dry-run", action="store_true", help="干跑:执行全部逻辑但回滚,不落库")
    g.add_argument("--commit", action="store_true", help="落库提交")
    args = ap.parse_args()
    asyncio.run(_execute(dry_run=args.dry_run))


if __name__ == "__main__":
    main()
